"""Stage 1 — turn whatever the client sent into the canonical schema.

Every client formats their spreadsheet differently. If the operator has to personally
interpret each one, this stays a service rather than a product, so intake is where that
problem is fought and won. Two nodes, because the spec's three "doors" into the schema
are really one flow with an optional model in the middle:

    ArkSheetProbe   the client's file  ->  columns, sample rows, a heuristic mapping
          |
          |  (optional) a vision/text LLM refines the mapping   <- door 3, auto-detect
          |  (optional) a saved per-client mapping is used      <- door 2, mapping
          |  (optional) nothing at all: the sheet is canonical  <- door 1, template
          v
    ArkVariationIntake   raw rows + mapping  ->  VARIANTS / SPECS / PRODUCT + validation

The probe emits a heuristic mapping of its own, so doors 1 and 2 frequently need no
model at all — a sheet whose headers already say `filename` and `axis:finish` maps
itself, and one that says `meta:attribute_pa_finish` is recognised by convention.

Structural quirks handled here rather than downstream, because a downstream node that
has to know about them is a downstream node that has been contaminated by one client's
layout: blank spacer rows between blocks, values populated only on a block's first row,
inconsistent casing, and state encoded as cell background colour.
"""

from __future__ import annotations

import csv
import io
import json
import os

from .schema import (
    AXIS_PREFIX,
    ValidationError,
    axis_names,
    canonical,
    normalise_hex,
    specs_index,
    spec_type_of,
    stable_json,
    validate_intake,
)

# Header prefixes a client sheet may already be using. WooCommerce variation exports
# name attribute columns `meta:attribute_pa_{taxonomy}`, so a sheet prepared for a store
# import is already declaring its own axes — recognising that convention means the
# reference client's sheet maps itself with no human involvement at all.
WOO_ATTRIBUTE_PREFIX = "meta:attribute_pa_"

_FILENAME_HINTS = ("image url", "image_url", "filename", "file name", "image",
                   "file", "output", "image name")
_STATUS_HINTS = ("status", "state", "done", "complete")
_REF_HINTS = ("reference image", "reference_image", "ref image", "ref_url",
              "reference url", "swatch", "sample image")
_HEX_HINTS = ("hex", "colour code", "color code", "hex code", "colour hex", "color hex")


def _sniff_encoding(path):
    """Work out how a client's export is actually encoded, from its first bytes.

    Never assume UTF-8. Excel's "Unicode Text" export is UTF-16 with a BOM, Google
    Sheets writes UTF-8 with a BOM, and a sheet saved on a Windows machine in a
    non-English locale is frequently cp1252. Guessing wrong produces
    `'utf-8' codec can't decode byte 0xff in position 0`, which tells the operator
    nothing about the file they actually have.
    """
    with open(path, "rb") as handle:
        head = handle.read(4)

    if head.startswith(b"\xff\xfe\x00\x00"):
        return "utf-32"
    if head.startswith(b"\x00\x00\xfe\xff"):
        return "utf-32"
    if head.startswith(b"\xff\xfe") or head.startswith(b"\xfe\xff"):
        return "utf-16"                      # the codec reads the BOM and picks the order
    if head.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"

    # No BOM. Try the two that matter, strictly, and take the first that parses.
    for candidate in ("utf-8-sig", "cp1252"):
        try:
            with open(path, "r", encoding=candidate) as handle:
                handle.read()
            return candidate
        except (UnicodeDecodeError, LookupError):
            continue
    return "latin-1"                          # decodes any byte; last resort, never fails


def _read_rows(path):
    """Read a client file into a list of dicts, preserving the original headers.

    CSV, TSV and JSON are handled directly. XLSX is read when openpyxl happens to be
    installed and otherwise reports the one-line fix rather than failing obscurely —
    every spreadsheet tool exports CSV, and a Google Sheet (the reference case) does so
    natively.
    """
    path = str(path or "").strip().strip('"')
    if not path:
        raise ValidationError("No sheet path given.")
    if not os.path.isfile(path):
        raise ValidationError("Sheet not found: %s" % path)

    ext = os.path.splitext(path)[1].lower()

    encoding = _sniff_encoding(path)

    if ext in (".json",):
        with open(path, "r", encoding=encoding) as handle:
            data = json.load(handle)
        if isinstance(data, dict):
            for key in ("rows", "variants", "data", "records"):
                if isinstance(data.get(key), list):
                    data = data[key]
                    break
        if not isinstance(data, list):
            raise ValidationError("%s is JSON but not a list of rows." % path)
        return [{str(k): v for k, v in row.items()} for row in data if isinstance(row, dict)]

    if ext in (".xlsx", ".xlsm"):
        try:
            import openpyxl
        except ImportError:
            raise ValidationError(
                "%s is an Excel file and openpyxl is not installed in this ComfyUI. "
                "Export the sheet as CSV and point this node at that instead "
                "(File > Download > Comma-separated values in Google Sheets)." % path)
        book = openpyxl.load_workbook(path, data_only=True)
        sheet = book[book.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h if h is not None else "") for h in rows[0]]
        out = []
        for raw in rows[1:]:
            out.append({headers[i]: ("" if v is None else v)
                        for i, v in enumerate(raw) if i < len(headers)})
        return out

    # A binary file reaching this point is almost always a spreadsheet saved with the
    # wrong extension. Say so, rather than letting the csv module produce nonsense.
    with open(path, "rb") as probe:
        head = probe.read(8)
    if head[:2] == b"PK":
        raise ValidationError(
            "%s is a ZIP-based file (an .xlsx/.ods saved with a .csv name). Re-export "
            "it as real CSV, or rename it to .xlsx so it is read as a workbook." % path)
    if head[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        raise ValidationError(
            "%s is a legacy .xls workbook. Open it and re-export as CSV." % path)

    # CSV / TSV / anything delimited. Sniff the delimiter so a tab-separated export
    # pasted out of a browser does not silently become one giant column.
    with open(path, "r", encoding=encoding, newline="") as handle:
        sample = handle.read(8192)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(handle, dialect=dialect)
        rows = [{str(k or ""): ("" if v is None else v) for k, v in row.items()}
                for row in reader]

    if encoding not in ("utf-8-sig",):
        print("[arkennemasis] %s decoded as %s" % (os.path.basename(path), encoding))
    return rows


def _is_blank_row(row) -> bool:
    """A spacer row between blocks. Dropped, never treated as a record."""
    return all(str(v or "").strip() == "" for v in row.values())


def _fill_down(rows, columns):
    """Carry a sparse column's last non-empty value forward.

    Specification data living inside a cell table is why these columns are sparse in the
    first place: the client wrote the finish's reference URL once, on the first row of
    that finish's block. Filling down restores the value to every row that inherits it.
    """
    carried = {}
    for row in rows:
        for column in columns:
            value = str(row.get(column) or "").strip()
            if value:
                carried[column] = value
            elif column in carried:
                row[column] = carried[column]
    return rows


def _suggest_mapping(rows):
    """A best-effort column mapping, from header conventions alone.

    Deliberately conservative: it proposes what it can recognise and stays silent about
    the rest, so that anything it is unsure of surfaces at the confirmation gate rather
    than being guessed at. A guess that looks right is worse than no guess.
    """
    if not rows:
        return {}, []
    headers = list(rows[0].keys())
    lower = {h: str(h or "").strip().lower() for h in headers}
    mapping = {"columns": {}, "fill_down": []}
    notes = []

    for header in headers:
        name = lower[header]
        if not name:
            continue

        if name.startswith(WOO_ATTRIBUTE_PREFIX):
            axis = canonical(name[len(WOO_ATTRIBUTE_PREFIX):])
            if axis:
                mapping["columns"][AXIS_PREFIX + axis] = header
                notes.append("'%s' -> axis '%s' (WooCommerce attribute convention)"
                             % (header, axis))
            continue

        if name.startswith(AXIS_PREFIX):
            mapping["columns"][name] = header
            notes.append("'%s' -> axis '%s' (already canonical)"
                         % (header, name[len(AXIS_PREFIX):]))
            continue

        if "filename" not in mapping["columns"] and name in _FILENAME_HINTS:
            mapping["columns"]["filename"] = header
            notes.append("'%s' -> filename (the primary key)" % header)
            continue

        if any(hint in name for hint in _REF_HINTS):
            mapping["columns"]["ref_url:?"] = header
            mapping["fill_down"].append(header)
            notes.append("'%s' looks like a reference-image column, but which axis it "
                         "specifies could not be determined — set 'ref_url:<axis>'."
                         % header)
            continue

        if any(hint in name for hint in _HEX_HINTS):
            mapping["columns"]["hex:?"] = header
            notes.append("'%s' looks like a hex column — set 'hex:<axis>'." % header)
            continue

        if "status" not in mapping["columns"] and name in _STATUS_HINTS:
            mapping["columns"]["status"] = header
            continue

    # Sparse columns are almost always block-scoped specification data.
    for header in headers:
        if header in mapping["fill_down"]:
            continue
        values = [str(r.get(header) or "").strip() for r in rows]
        filled = [v for v in values if v]
        if filled and len(filled) < len(values) * 0.5:
            mapping["fill_down"].append(header)
            notes.append("'%s' is populated on %d of %d rows — filling down."
                         % (header, len(filled), len(values)))

    return mapping, notes


class ArkSheetProbe:
    CATEGORY = "arkennemasis/Variation"
    FUNCTION = "run"
    RETURN_TYPES = ("STRING", "STRING", "STRING", "INT")
    RETURN_NAMES = ("raw_json", "columns_report", "suggested_mapping", "row_count")
    DESCRIPTION = (
        "Read a client's variation sheet without interpreting it. Emits the raw rows, a "
        "human/LLM-readable description of the columns, and a heuristic column mapping. "
        "Feed columns_report to an LLM to auto-detect the mapping, or hand-write one."
    )

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "sheet_path": ("STRING", {
                    "default": "",
                    "tooltip": "CSV, TSV or JSON export of the client's sheet. XLSX works "
                               "only if openpyxl is installed — export CSV instead.",
                }),
            },
            "optional": {
                "sample_rows": ("INT", {
                    "default": 8, "min": 1, "max": 100,
                    "tooltip": "How many example rows to include in columns_report. "
                               "Enough for a model to see the shape, not so many that "
                               "the whole sheet is pasted into a prompt.",
                }),
            },
        }

    @classmethod
    def IS_CHANGED(cls, sheet_path="", **kwargs):
        # Re-read when the file changes on disk. A client re-exporting their sheet must
        # not be served a cached view of the previous version.
        try:
            return os.path.getmtime(str(sheet_path).strip().strip('"'))
        except OSError:
            return float("nan")

    def run(self, sheet_path, sample_rows=8):
        rows = _read_rows(sheet_path)
        rows = [r for r in rows if not _is_blank_row(r)]
        if not rows:
            raise ValidationError("%s has no data rows." % sheet_path)

        mapping, notes = _suggest_mapping(rows)
        headers = list(rows[0].keys())

        lines = ["SHEET: %s" % os.path.basename(str(sheet_path)),
                 "ROWS (excluding blank spacer rows): %d" % len(rows),
                 "COLUMNS: %d" % len(headers), "",
                 "COLUMN DETAIL — header, populated count, distinct count, examples"]
        for header in headers:
            values = [str(r.get(header) or "").strip() for r in rows]
            filled = [v for v in values if v]
            distinct = sorted(set(filled))
            examples = ", ".join(repr(v) for v in distinct[:4])
            lines.append("  %-42s filled %4d/%d  distinct %4d  e.g. %s"
                         % (header[:42], len(filled), len(values), len(distinct), examples))

        lines += ["", "SAMPLE ROWS"]
        for row in rows[:int(sample_rows)]:
            lines.append("  " + json.dumps(row, ensure_ascii=False)[:400])

        if notes:
            lines += ["", "HEURISTIC OBSERVATIONS"]
            lines += ["  - " + n for n in notes]

        report = "\n".join(lines)
        print("[arkennemasis] sheet probe: %d rows, %d columns" % (len(rows), len(headers)))
        return (json.dumps(rows, ensure_ascii=False), report,
                json.dumps(mapping, indent=2, ensure_ascii=False), len(rows))


class ArkVariationIntake:
    CATEGORY = "arkennemasis/Variation"
    FUNCTION = "run"
    RETURN_TYPES = ("STRING", "STRING", "STRING", "BOOLEAN", "INT")
    RETURN_NAMES = ("intake_json", "report", "problems", "ok", "cell_count")
    DESCRIPTION = (
        "Normalise a client's sheet into the canonical VARIANTS / SPECS / PRODUCT "
        "tables and run every pre-generation validation. Nothing downstream sees the "
        "client's layout. A failure here costs nothing; the same failure at cell 380 "
        "costs 380 images."
    )

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "raw_json": ("STRING", {
                    "multiline": True, "default": "[]", "forceInput": True,
                    "tooltip": "Rows from Sheet Probe.",
                }),
                "mapping_json": ("STRING", {
                    "multiline": True, "default": "{}",
                    "tooltip": "Which client column means what. Use Sheet Probe's "
                               "suggestion, an LLM's proposal, or a saved per-client "
                               "mapping. Keys: columns{filename, axis:<name>, "
                               "ref_url:<axis>, hex:<axis>, status}, fill_down[], "
                               "product, product_display, naming_pattern, plates[].",
                }),
            },
            "optional": {
                "specs_path": ("STRING", {
                    "default": "",
                    "tooltip": "Optional separate SPECS file (CSV/JSON): one row per "
                               "axis value with axis, value, hex, ref_url, display, "
                               "filename_token, description. Merged over anything "
                               "lifted out of the variant sheet.",
                }),
                "product_id": ("STRING", {
                    "default": "",
                    "tooltip": "Canonical product id. Overrides the mapping's value. "
                               "Blank falls back to the mapping, then the sheet.",
                }),
                "naming_pattern": ("STRING", {
                    "default": "",
                    "tooltip": "Filename template, e.g. "
                               "'{axis-a.filename_token}-in-{axis-b.filename_token}.png'. "
                               "Blank = take the client's filenames exactly as given, "
                               "which is the safer default.",
                }),
                "strict": ("BOOLEAN", {
                    "default": True,
                    "tooltip": "On: any validation problem blocks the run. Off: problems "
                               "are reported and the run continues — for inspecting a "
                               "messy sheet, never for a real production run.",
                }),
                "write_to": ("STRING", {
                    "default": "",
                    "tooltip": "Optional folder to write intake.json into, so the "
                               "orchestrator and later runs can read it without "
                               "re-parsing the client's file.",
                }),
            },
        }

    def run(self, raw_json, mapping_json, specs_path="", product_id="",
            naming_pattern="", strict=True, write_to=""):
        try:
            rows = json.loads(raw_json or "[]")
        except ValueError as exc:
            raise ValidationError("raw_json is not valid JSON: %s" % exc)
        try:
            mapping = json.loads(mapping_json or "{}")
        except ValueError as exc:
            raise ValidationError("mapping_json is not valid JSON: %s" % exc)

        columns = mapping.get("columns") or {}
        unresolved = [k for k in columns if k.endswith(":?")]
        if unresolved:
            raise ValidationError(
                "The mapping still has unresolved placeholders: %s. Replace the '?' "
                "with the axis name each column specifies." % ", ".join(unresolved))

        rows = [r for r in rows if isinstance(r, dict) and not _is_blank_row(r)]
        if not rows:
            raise ValidationError("No data rows after dropping blank spacer rows.")

        fill_down = [c for c in (mapping.get("fill_down") or []) if c in rows[0]]
        if fill_down:
            rows = _fill_down(rows, fill_down)

        # ── VARIANTS ────────────────────────────────────────────────────────
        axis_columns = {k[len(AXIS_PREFIX):]: v for k, v in columns.items()
                        if k.startswith(AXIS_PREFIX)}
        if not axis_columns:
            raise ValidationError(
                "The mapping declares no axis columns. At least one "
                "'%s<name>' entry is required." % AXIS_PREFIX)

        filename_column = columns.get("filename")
        status_column = columns.get("status")

        variants = []
        for row in rows:
            record = {}
            for axis, column in axis_columns.items():
                record[AXIS_PREFIX + canonical(axis)] = canonical(row.get(column))
            if filename_column:
                record["filename"] = str(row.get(filename_column) or "").strip()
            if status_column:
                # Cell background colour encoding state is invisible to any script
                # reading values, so whatever explicit status column exists is lifted
                # here and colour stops being load-bearing from this point on.
                record["status"] = str(row.get(status_column) or "").strip().lower()
            variants.append(record)

        # ── SPECS lifted out of the variant sheet ───────────────────────────
        # This is the normalisation the spec calls "two tables, not one": specification
        # data living inside a cell table is pulled out into its own table, keyed by
        # axis value, so 81 rows collapse to 18 specifications.
        lifted = {}
        for key, column in columns.items():
            for prefix, field in (("ref_url:", "ref_url"), ("hex:", "hex"),
                                  ("description:", "description"),
                                  ("display:", "display"),
                                  ("filename_token:", "filename_token")):
                if key.startswith(prefix):
                    axis = canonical(key[len(prefix):])
                    axis_column = axis_columns.get(axis)
                    if not axis_column:
                        raise ValidationError(
                            "Mapping has '%s' but no '%s%s' column to key it against."
                            % (key, AXIS_PREFIX, axis))
                    for row in rows:
                        value = canonical(row.get(axis_column))
                        payload = str(row.get(column) or "").strip()
                        if not value or not payload:
                            continue
                        entry = lifted.setdefault((axis, value),
                                                  {"axis": axis, "value": value})
                        entry.setdefault(field, payload)

        specs = list(lifted.values())

        # Inline specs from the mapping, then a separate specs file — later sources win
        # on a field-by-field basis so a hand-corrected hex overrides a lifted one.
        for entry in (mapping.get("specs") or []):
            if isinstance(entry, dict):
                specs.append(entry)
        if str(specs_path or "").strip():
            for entry in _read_rows(specs_path):
                specs.append(entry)

        # ── PRODUCT ─────────────────────────────────────────────────────────
        product = canonical(product_id) or canonical(mapping.get("product"))
        if not product:
            raise ValidationError(
                "No product id, so there is nothing to key this run's files and job "
                "records on.\n\n"
                "FIX EITHER ONE:\n"
                "  1. Type an id into this node's 'product_id' widget, e.g. "
                "'acme-widget-01'; or\n"
                "  2. Add a \"product\" key to mapping_json:\n"
                "       { \"product\": \"acme-widget-01\", \"columns\": { ... } }\n\n"
                "If you are running one of the bundled demos, paste that demo's whole "
                "mapping.json into mapping_json — it already carries the product id.")
        plates = mapping.get("plates") or [{"id": "front", "file": ""}]
        if not isinstance(plates, list) or not plates:
            raise ValidationError("mapping.plates must be a non-empty list.")

        pattern = str(naming_pattern or mapping.get("naming_pattern") or "").strip()

        product_row = {
            "product": product,
            "display_name": str(mapping.get("product_display") or "").strip() or product,
            "plates": plates,
            "naming_pattern": pattern,
            "colour_profile": str(mapping.get("colour_profile") or "").strip() or "sRGB",
            "notes": str(mapping.get("notes") or ""),
        }

        # ── Validation ──────────────────────────────────────────────────────
        problems = validate_intake(variants, specs, product_row, plates)

        # Parse-back: the client's filename must agree with the client's axis columns.
        # Disagreement is a hard error rather than a warning — it means either the
        # filename or the attribute columns are wrong, and shipping under either
        # assumption puts a mislabelled image on a product page.
        index = specs_index(specs)
        checked = matched = 0
        for i, row in enumerate(variants, start=1):
            name = str(row.get("filename") or "")
            if not name:
                continue
            checked += 1
            bad = []
            for axis in axis_columns:
                axis_c = canonical(axis)
                value = row.get(AXIS_PREFIX + axis_c)
                entry = index.get((axis_c, value))
                token = (entry or {}).get("filename_token") or value
                if token and token not in name.lower():
                    bad.append((axis_c, value, token))
            if bad:
                detail = "; ".join("axis '%s' = '%s' (expected token '%s')" % b for b in bad)
                problems.append(
                    "VARIANTS row %d filename '%s' does not contain %s. Either the "
                    "filename or the attribute column is wrong — set filename_token in "
                    "SPECS if the filename legitimately uses a shorter form."
                    % (i, name, detail))
            else:
                matched += 1

        cell_count = len(variants) * max(1, len(plates))
        ok = not problems

        lines = [
            "INTAKE — %s" % product_row["display_name"],
            "  product        : %s" % product,
            "  plates         : %d (%s)" % (len(plates),
                                            ", ".join(str(p.get("id")) for p in plates)),
            "  axes           : %d (%s)" % (len(axis_columns),
                                            ", ".join(sorted(axis_columns))),
            "  variant rows   : %d" % len(variants),
            "  specifications : %d" % len(index),
            "  cells (rows x plates): %d" % cell_count,
            "  naming pattern : %s" % (pattern or "(client filenames used verbatim)"),
            "  filename check : %d/%d rows agree with their axis columns"
            % (matched, checked),
        ]
        if fill_down:
            lines.append("  filled down    : %s" % ", ".join(fill_down))

        by_axis = {}
        for (axis, value), entry in sorted(index.items()):
            by_axis.setdefault(axis, []).append((value, spec_type_of(entry)))
        lines.append("")
        lines.append("SPECIFICATIONS BY AXIS")
        for axis, values in sorted(by_axis.items()):
            kinds = {}
            for _v, kind in values:
                kinds[kind] = kinds.get(kind, 0) + 1
            lines.append("  %-24s %2d values  %s"
                         % (axis, len(values),
                            ", ".join("%s x%d" % (k, n) for k, n in sorted(kinds.items()))))
            for value, kind in values:
                lines.append("      %-30s %s" % (value, kind))

        if problems:
            lines += ["", "PROBLEMS (%d)" % len(problems)]
            lines += ["  %d. %s" % (i, p) for i, p in enumerate(problems, start=1)]
        else:
            lines += ["", "All intake validations passed."]

        report = "\n".join(lines)

        intake = {
            "schema_version": "1.0",
            "product": product_row,
            "variants": variants,
            "specs": [dict(v) for v in index.values()],
            "axes": sorted(axis_columns),
            "cell_count": cell_count,
            "ok": ok,
            "problems": problems,
        }
        intake["intake_hash"] = __import__("hashlib").sha256(
            stable_json({k: intake[k] for k in ("product", "variants", "specs")})
            .encode("utf-8")).hexdigest()

        destination = str(write_to or "").strip()
        if destination:
            if not os.path.isabs(destination):
                try:
                    import folder_paths
                    destination = os.path.join(folder_paths.get_output_directory(),
                                               destination)
                except Exception:
                    destination = os.path.abspath(destination)
            os.makedirs(destination, exist_ok=True)
            path = os.path.join(destination, "intake.json")
            with open(path, "w", encoding="utf-8") as handle:
                json.dump(intake, handle, indent=2, ensure_ascii=False)
            print("[arkennemasis] intake -> %s" % path)

        print("[arkennemasis] intake: %d variants, %d specs, %d cells, %s"
              % (len(variants), len(index), cell_count,
                 "OK" if ok else "%d PROBLEM(S)" % len(problems)))

        if problems and strict:
            raise ValidationError(
                "Intake found %d problem(s). Fix the sheet or the mapping, or turn "
                "'strict' off to inspect.\n\n%s"
                % (len(problems), "\n".join("  - " + p for p in problems)))

        return (json.dumps(intake, ensure_ascii=False), report,
                "\n".join(problems), ok, cell_count)


NODE_CLASS_MAPPINGS = {
    "ArkSheetProbe": ArkSheetProbe,
    "ArkVariationIntake": ArkVariationIntake,
}

NODE_DISPLAY_NAME_MAPPINGS = {
    "ArkSheetProbe": "arkennemasis Sheet Probe (read a client sheet)",
    "ArkVariationIntake": "arkennemasis Variation Intake (normalise + validate)",
}
